import pandas as pd
import openpyxl
import json
from pathlib import Path
import os
from dotenv import load_dotenv
from azure.identity import DefaultAzureCredential
from openai import AzureOpenAI


def process_merged_headers(file_path: str, output_json_path: str = None) -> pd.DataFrame:
    """
    Read an Excel file with merged cells in first 2 rows, concatenate with row 3
    to create single header row, and return as DataFrame.
    
    Args:
        file_path: Path to the Excel file
        output_json_path: Optional path to save JSON output
        
    Returns:
        pandas DataFrame with concatenated headers
    """
    # Load workbook with openpyxl to handle merged cells
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get all merged cell ranges in the first 2 rows
    merged_ranges = [mr for mr in ws.merged_cells.ranges if mr.min_row <= 2]
    
    # Create a mapping of column index to merged cell values
    merged_cell_values = {}
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        
        # Get the value from the top-left cell of the merged range
        cell_value = ws.cell(min_row, min_col).value
        
        if cell_value:
            # Store the merged value for all columns it spans
            for col in range(min_col, max_col + 1):
                if col not in merged_cell_values:
                    merged_cell_values[col] = []
                merged_cell_values[col].append(str(cell_value))
    
    # Build the final header row by concatenating merged values with row 3
    num_cols = ws.max_column
    final_headers = []
    
    for col in range(1, num_cols + 1):
        # Get values from merged cells in rows 1-2
        prefix_parts = merged_cell_values.get(col, [])
        
        # Get value from row 3
        row3_value = ws.cell(3, col).value
        row3_str = str(row3_value) if row3_value else ""
        
        # Concatenate all parts
        all_parts = prefix_parts + [row3_str] if row3_str else prefix_parts
        
        # Join with separator
        final_header = " - ".join([part for part in all_parts if part])
        final_headers.append(final_header if final_header else f"Column_{col}")
    
    # Read the data starting from row 4 (after the 3 header rows)
    df = pd.read_excel(file_path, header=None, skiprows=3)
    
    # Trim to actual column count
    df = df.iloc[:, :len(final_headers)]
    
    # Assign the concatenated headers
    df.columns = final_headers
    
    # Remove completely empty rows and columns
    df = df.dropna(axis=0, how='all')
    df = df.dropna(axis=1, how='all')
    
    return df


def extract_brands_with_llm(text: str, client: AzureOpenAI) -> str:
    """
    Use LLM to extract brand names from image filename patterns.
    
    Args:
        text: Text containing image filenames like [BandL_Spring22_Kmart_W120.png]
        client: Azure OpenAI client
        
    Returns:
        Text with brand names extracted and cleaned
    """
    if pd.isna(text) or text == "" or "[" not in str(text):
        return text
    
    prompt = f"""Extract and replace image filenames with their brand names.

Rules:
- [BandL_Spring22_Kmart_W120.png] -> Kmart
- [Target_(120px).png] -> Target
- [Best_Less_(120_x_40).png] -> Best_Less
- If multiple brands in a comma-separated list, replace each one
- Keep all other text unchanged

Text: {text}

Output only the cleaned text with brand names:"""

    try:
        response = client.chat.completions.create(
            model=os.getenv("AZURE_OPENAI_CHATGPT_DEPLOYMENT"),
            messages=[
                {"role": "system", "content": "You extract brand names from image filenames. Return only the cleaned text."},
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            max_tokens=500
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"[warning] LLM extraction failed for text, using original: {e}")
        return text


def clean_df(df: pd.DataFrame, use_llm: bool = False) -> pd.DataFrame:
    """
    Clean the dataframe by:
    - Converting dates to standard format
    - Replacing image filename patterns with brand names
    
    Args:
        df: Input DataFrame
        use_llm: If True, use LLM for brand extraction; otherwise use regex
        
    Returns:
        Cleaned DataFrame
    """
    # Make a copy to avoid modifying original
    df = df.copy()
    
    # Define image replacements (used for both column names and values)
    image_replacements = {
        # Kmart patterns
        r'\[BandL_Spring22_Kmart_W120\.(png|jpg)\]': 'Kmart',
        r'\[Kmart_\(120px\)\.(png|jpg)\]': 'Kmart',
        
        # Target patterns
        r'\[BandL_Spring22_Target_W120\.(png|jpg)\]': 'Target',
        r'\[Target_\(120px\)\.(png|jpg)\]': 'Target',
        
        # Best & Less patterns
        r'\[BandL_Spring22_BL_W120\.(png|jpg)\]': 'Best_Less',
        r'\[Best_Less_\(120_x_40\)\.(png|jpg)\]': 'Best_Less',
        
        # Big W patterns
        r'\[BandL_Spring22_BigW_W120\.(png|jpg)\]': 'Big_W',
        r'\[Big_W_\(85px\)\.(png|jpg)\]': 'Big_W',
        
        # Cotton On patterns
        r'\[BandL_Spring22_CottonOn_W120\.(png|jpg)\]': 'Cotton_On',
        r'\[Cotton_On_\(120px\)\.(png|jpg)\]': 'Cotton_On',
    }
    
    # Convert date column
    if "Submitted Date" in df.columns:
        df["Submitted Date"] = pd.to_datetime(df["Submitted Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    
    # First, replace patterns in column names
    import re
    new_columns = []
    for col in df.columns:
        new_col = str(col)
        for pattern, replacement in image_replacements.items():
            new_col = re.sub(pattern, replacement, new_col)
        new_columns.append(new_col)
    df.columns = new_columns
    
    if use_llm:
        # Initialize Azure OpenAI client
        credential = DefaultAzureCredential()
        token_provider = lambda: credential.get_token("https://cognitiveservices.azure.com/.default").token
        
        client = AzureOpenAI(
            api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"),
            azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
            azure_ad_token_provider=token_provider
        )
        
        # Apply LLM-based extraction to all string columns
        for idx, col in enumerate(df.columns):
            col_data = df.iloc[:, idx]
            if col_data.dtype == 'object':
                #print(f"[info] Processing column '{col}' with LLM...")
                df.iloc[:, idx] = col_data.apply(lambda x: extract_brands_with_llm(x, client) if pd.notna(x) else x)
    else:
        # Use regex patterns on column values
        for idx, col in enumerate(df.columns):
            col_data = df.iloc[:, idx]
            if col_data.dtype == 'object':
                for pattern, replacement in image_replacements.items():
                    df.iloc[:, idx] = col_data.astype(str).str.replace(pattern, replacement, regex=True)
                    col_data = df.iloc[:, idx]  # Update col_data for next pattern
    
    return df


if __name__ == "__main__":
    # Load environment variables
    load_dotenv(".azure/Personas/.env")
    
    # Example usage
    input_file = "data_prep/input_datasets/BestLessBrandTrackingWinter2024_v2.xlsx"
    output_json = "data_prep/input_datasets/structured_output.json"
    
    try:
        # Process merged headers
        df = process_merged_headers(input_file)
        print(f"[ok] Loaded {len(df)} rows with {len(df.columns)} columns")
        
        # Clean data - set use_llm=True to use LLM, False for regex
        df = clean_df(df, use_llm=False)  # Start with regex for speed
        print(f"[ok] Cleaned {len(df)} rows")
        
        # Set Come-back-Later Code as index to remove it from columns
        if "Come-back-Later Code" in df.columns:
            df = df.set_index("Come-back-Later Code")
            print(f"[ok] Set Come-back-Later Code as index")
        
        # Build output JSON structure
        output = []
        for idx, row in df.iterrows():
            # idx is now the Come-back-Later Code value
            comeback_code = str(idx) if pd.notna(idx) else None
            
            # Extract year from Submitted Date
            submitted_date = row.get("Submitted Date") if isinstance(row.get("Submitted Date"), str) else None
            year = None
            if submitted_date:
                try:
                    year = str(pd.to_datetime(submitted_date).year)
                except:
                    year = None
            
            # Build description - use enumerate to handle duplicate column names
            desc_parts = []
            for col_idx, col_name in enumerate(df.columns.tolist()):
                if col_name == "Complete Type":
                    continue
                # Use iloc to access by position to avoid duplicate column issues
                value = row.iloc[col_idx]
                # Check if value is not None/NaN and not empty string
                if pd.notna(value) and str(value).strip() != "":
                    desc_parts.append(f"{col_name}: {value}")
            
            # Get state value - use iloc for first occurrence of "Complete Type"
            complete_type_idx = df.columns.tolist().index("Complete Type") if "Complete Type" in df.columns.tolist() else None
            state = row.iloc[complete_type_idx] if complete_type_idx is not None else None
            
            obj = {
                "AreaPath": "B&LPersonas",
                "AssignedTo": None,
                "Categories": year,
                "ChangedDate": None,
                "ClosedDate": None,
                "CreatedDate": submitted_date,
                "Description": "\n".join(desc_parts),
                "Id": comeback_code,
                "State": state,
                "StateChangeDate": None,
                "Tags": "B&LSurvey",
                "Title": comeback_code
            }
            output.append(obj)
        
        # Save output JSON as a list
        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        
        print(f"[ok] Saved JSON output to {output_json}")
        print(f"[ok] Generated {len(output)} records")
        print(f"[ok] Sample record: {json.dumps(output[0], indent=2)}")
    except FileNotFoundError:
        print(f"[error] File not found: {input_file}")
    except Exception as e:
        print(f"[error] Failed to process file: {e}")
        import traceback
        traceback.print_exc()